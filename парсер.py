import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter

workbook = openpyxl.load_workbook('parser_data.xlsx')

sheet = workbook.get_sheet_by_name('Games Data')

sheet.delete_rows(1,sheet.max_row)
sheet.delete_cols(1,sheet.max_column)

number = 0

# Отправляем GET-запрос к странице, которую хотим спарсить
for i in range(1, 83+1):	
	url = 'https://itorrents-igruha.org/igri-dly-slabih-pc/page/' + str(i) + '/'

	response = requests.get(url)

# Создаем объект BeautifulSoup из ответа сервера
	soup = BeautifulSoup(response.content, 'lxml')

# Находим нужные элементы на странице
	all_imgs = soup.find_all('div', {'class':'article-film-image'})

	
	for i in all_imgs:
	

		a_tags = i.find('a')
		games_titles = a_tags.get('title')

		games_hrefs = a_tags.get('href')

		if games_titles != "Важная информация!":
			number = number + 1	
			games_rows = [number, games_titles, games_hrefs]
		
			sheet.append(games_rows)



# Автоматическая настройка ширины столбцов
for column in sheet.columns:
    column_name = get_column_letter(column[0].column)
    max_length = 0
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    sheet.column_dimensions[column_name].width = adjusted_width

# Сохранение файла
workbook.save('parser_data.xlsx')